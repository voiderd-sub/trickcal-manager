import openpyxl
import numpy as np
from math import comb, log
from scipy.optimize import minimize, approx_fprime
from scipy.stats import norm

def prob_S_given_N(s, N, p0, p1, p2):
    """P(S=s | N) 계산"""
    total_prob = 0.0
    for m in range(0, s // 2 + 1):
        k2 = m          # 2가 나온 횟수
        k1 = s - 2*m    # 1이 나온 횟수
        k0 = N - k1 - k2
        if k0 < 0:
            continue
        coeff = comb(N, k0) * comb(N - k0, k1)  # 다항계수 N!/(k0! k1! k2!)
        total_prob += coeff * (p0**k0) * (p1**k1) * (p2**k2)
    return total_prob

def neg_log_likelihood(params, N_list, S_list):
    """(p0, p1)를 받아서 음의 로그우도 반환"""
    p0, p1 = params
    p2 = 1 - p0 - p1
    if p0 <= 0 or p1 <= 0 or p2 <= 0:
        return np.inf
    ll = 0.0
    for N, S in zip(N_list, S_list):
        prob = prob_S_given_N(S, N, p0, p1, p2)
        if prob <= 0:
            return np.inf
        ll += log(prob)
    return -ll

def _calculate_hessian(p_mle, N_list, S_list):
    """음의 로그우도 함수의 헤세 행렬을 수치적으로 계산합니다."""
    n_params = len(p_mle)
    hessian = np.zeros((n_params, n_params))
    
    # approx_fprime에 넘겨주기 위해 p만을 인자로 받는 함수를 정의합니다.
    func = lambda p: neg_log_likelihood(p, N_list, S_list)
    
    # 중앙차분을 이용해 그래디언트의 야코비안(헤세 행렬)을 계산합니다.
    epsilon = np.sqrt(np.finfo(float).eps)
    for i in range(n_params):
        p_plus_h = p_mle.copy()
        p_plus_h[i] += epsilon
        grad_plus_h = approx_fprime(p_plus_h, func, epsilon)

        p_minus_h = p_mle.copy()
        p_minus_h[i] -= epsilon
        grad_minus_h = approx_fprime(p_minus_h, func, epsilon)

        hessian[:, i] = (grad_plus_h - grad_minus_h) / (2 * epsilon)
    
    # 계산된 헤세 행렬이 대칭행렬이 되도록 보정합니다.
    return (hessian + hessian.T) / 2

def estimate_p_full_mle(filepath, sheetname=None, alpha=0.05):
    # --------------------------
    # 1. 데이터 읽기
    # --------------------------
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheetname] if sheetname else wb.active
    
    N_list, S_list = [], []
    row = 3
    while True:
        N_val = ws.cell(row=row, column=2).value  # B열
        S_val = ws.cell(row=row, column=6).value  # F열
        if N_val is None or S_val is None:
            break
        N_list.append(int(N_val))
        S_list.append(int(S_val))
        row += 1
    
    # --------------------------
    # 2. 최적화
    # --------------------------
    init = [0.95, 0.025]  # 초기값 (p0, p1)
    bounds = [(1e-6, 1-1e-6), (1e-6, 1-1e-6)]  # 경계조건
    cons = ({'type': 'ineq', 'fun': lambda x: 1 - x[0] - x[1] - 1e-6})  # p2 >= 0
    
    res = minimize(neg_log_likelihood, init, args=(N_list, S_list),
                   bounds=bounds, constraints=cons, method='SLSQP')
    
    if not res.success:
        raise RuntimeError("최적화 실패: " + res.message)
    
    p0, p1 = res.x
    p2 = 1 - p0 - p1

    # --------------------------
    # 3. 신뢰구간 계산
    # --------------------------
    # 헤세 행렬 계산
    hessian_mat = _calculate_hessian(res.x, N_list, S_list)

    # 공분산 행렬 계산 (헤세 행렬의 역행렬)
    try:
        cov_mat = np.linalg.inv(hessian_mat)
    except np.linalg.LinAlgError:
        print("경고: 헤세 행렬이 특이 행렬(singular)이므로 신뢰구간을 계산할 수 없습니다.")
        return p0, p1, p2, None, None, None

    # 표준오차 계산
    se_p0 = np.sqrt(cov_mat[0, 0]) if cov_mat[0, 0] >= 0 else np.nan
    se_p1 = np.sqrt(cov_mat[1, 1]) if cov_mat[1, 1] >= 0 else np.nan
    
    # p2 = 1 - p0 - p1 이므로, 오차 전파 사용
    var_p2 = cov_mat[0, 0] + cov_mat[1, 1] + 2 * cov_mat[0, 1]
    se_p2 = np.sqrt(var_p2) if var_p2 >= 0 else np.nan

    # 신뢰구간 계산
    z_score = norm.ppf(1 - alpha / 2)
    ci_p0 = (p0 - z_score * se_p0, p0 + z_score * se_p0)
    ci_p1 = (p1 - z_score * se_p1, p1 + z_score * se_p1)
    ci_p2 = (p2 - z_score * se_p2, p2 + z_score * se_p2)

    return p0, p1, p2, ci_p0, ci_p1, ci_p2

# 사용 예시
if __name__ == "__main__":
    filepath = "C:/Users/USER/Downloads/trickcal/도안 드랍테이블.xlsx"
    sheetname = "크레용 12지 (내 데이터만)"
    p0, p1, p2, ci_p0, ci_p1, ci_p2 = estimate_p_full_mle(filepath, sheetname)

    print(f"p0_hat = {p0:.4f}", end="")
    if ci_p0:
        print(f" (95% CI: {ci_p0[0]:.4f}, {ci_p0[1]:.4f})")
    else:
        print()

    print(f"p1_hat = {p1:.4f}", end="")
    if ci_p1:
        print(f" (95% CI: {ci_p1[0]:.4f}, {ci_p1[1]:.4f})")
    else:
        print()

    print(f"p2_hat = {p2:.4f}", end="")
    if ci_p2:
        print(f" (95% CI: {ci_p2[0]:.4f}, {ci_p2[1]:.4f})")
    else:
        print()


    print(f"mean \sum x = {p1 + p2 * 2:.4f}")